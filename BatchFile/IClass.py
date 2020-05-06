import os
import pandas as pd
import shutil
import datetime
import zipfile as zf
import re
import openpyxl

# column name of input file
FILE_TYPE = "file type"
DB_NAME = "db name"
FILE_EXTENSION = "file extension"
SOURCE_PATH = "source path"
DESTINATION_PATH = "destination path"
HAS_RENAME = "has rename"
RENAME_TO = "rename to"
HAS_PASSWORD = "has password"
PASSWORD_TO = "password to"
MEDIATOR_PATH = "mediator path"
HAS_ZIP = "has zip"
SIZE = "size"

# column name of log file
L_MOVED = 'moved'
L_DB_NAME = "db name"
L_SIZE = "file size status"
L_SIZE_CHANGE = "file size difference (%)(KB)"
L_RENAME = "renamed"
L_ZIPPED = "zipped"
L_DATE_STATUS = "lastest date status"
L_DATE = "date"
L_SUMMARY = "summary"

class IToLog:
    def __init__(self, pathExcelFile):
        self.workSheet = None
        self.pathFile = pathExcelFile
        self.row = 1
        try:
            self.workBook = openpyxl.load_workbook(self.pathFile)
        except (Exception,):  # if work book does't exist then create one
            self.workBook = openpyxl.workbook.Workbook()

    def selectWorkSheet(self, sheetName="Sheet1"):
        """this method will delete the old data of the workbook"""
        if sheetName in self.workBook.sheetnames:
            self.workBook.remove(self.workBook[sheetName])
            self.workSheet = self.workBook.create_sheet(sheetName)
        else:
            self.workSheet = self.workBook.create_sheet(sheetName)

    def setColumn(self, columns:list):
        for index, name in columns:
            self.workSheet.cell(column=index+1, row=self.row).value = name
        self.row += 1

    def writeCell(self, data:list):
        for index, dateInCell in data:
            self.workSheet.cell(column=index + 1, row=self.row).value = dateInCell
        self.row += 1

    def saveAction(self):
        self.workBook.save(self.pathFile)

    def closeWorkBook(self):
        self.workBook.close()

class IZipFile:
    def __init__(self, zipFilePath):
        self.zipFile = zf.ZipFile(zipFilePath)
        self.zipFilePath = zipFilePath

    def findFileInZipFile(self, extension):
        fileExtension = ""
        for file in self.zipFile.infolist():
            if file.filename.endswith(extension.lower()):
                fileExtension = file  # assign fileTxt to file object
                break
            else:
                continue
        if fileExtension != "":
            path = self.zipFilePath + "\\" + fileExtension.filename
            return path
        else:
            return ""

    @staticmethod
    def toZipFile(zipNamePath, filePathList, password=None):
        try:
            with zf.ZipFile(zipNamePath + ".zip", "w") as zipObj:
                for filePath in filePathList:
                    zipObj.write(filePath)
                if password is not None:
                    zipObj.setpassword(password)
                zipObj.close()
        except Exception:
            return False
        else:
            return True

    @staticmethod
    def checkZipFileDate(zipFilePath):
        try:
            t = os.path.getmtime(zipFilePath)
        except (FileNotFoundError, Exception):
            return ""
        else:
            modified_date = datetime.datetime.fromtimestamp(t)
            modified_date = str(modified_date)
            return modified_date

    def getNumberOfFiles(self):
        """
        this method get the number of files in the zip file
        :return:
        """
        size = len(self.zipFile.infolist())
        return size

    def getFileDateInZipFile(self):
        fileDateDict = {}
        for file in self.zipFile.infolist():
            fileName = file.filename
            filePath = self.zipFilePath + "\\" + fileName
            fileLastModified = os.path.getmtime(filePath)
            fileDateDict[fileName] = fileLastModified
        return fileDateDict

class IFile:
    def __init__(self):
        pass

    @staticmethod
    def readUserInputFile(pathExcel):
        userInput = pd.DataFrame(pd.read_excel(pathExcel))
        # list of array
        dbNameList = userInput[DB_NAME].unique().tolist()
        # group by db name
        userInput = userInput.groupby(by=DB_NAME)
        for dbName in dbNameList:
            thatDbNameDataFrame = pd.DataFrame(userInput.get_group(dbName[0]))
            # take values out of the data frame as DICTIONARY, by making "key" as string
            # of the user-excel column, "value" is list of value corresponding to that column
            # which is each "value" is the same size of the number of destination path
            dictUserInput = thatDbNameDataFrame.to_dict("index")
            size = len(thatDbNameDataFrame)
            # this size is the number of "value" of each "key" in DICTIONARY
            dictUserInput[SIZE] = size
            # deal with nan value at caller
            yield dictUserInput


    @staticmethod
    def getFileDate(pathFile):
        try:
            t = os.path.getmtime(pathFile)
        except (FileNotFoundError, Exception):
            return ""
        else:
            modified_date = datetime.datetime.fromtimestamp(t)
            modified_date = str(modified_date)
            return modified_date

    @staticmethod
    def copyWithMetaData(sourceDirectory, destinationDirectory):
        try:
            shutil.copy2(sourceDirectory, destinationDirectory)
        except (FileNotFoundError, Exception):
            return False
        else:
            return True

    @staticmethod
    def renameFile(oldFileNamePath, newFileName):
        try:
            newFileNamePath = os.path.dirname(oldFileNamePath) + "\\" + newFileName
            os.rename(oldFileNamePath, newFileNamePath)
        except (FileExistsError, Exception):
            return False
        else:
            return True

    @staticmethod
    def getFileSize(filePath):
        try:
            size = os.path.getsize(filePath)
        except (FileNotFoundError, Exception) :
            return 0
        else:
            return size

    @staticmethod
    def splitExtension(ExtensionString: str, separator="/"):
        """
        this method split the string of extension
        as such ".CHR/.HED" to the list. So for this sample,
        the list will be "[".CHR", ".HED"]"
        """
        ExtensionList = ExtensionString.split(separator)
        return ExtensionList

    @staticmethod
    def findFile(sourcePath, fileName, extensionList:list):
        """
        this method find file's path in the specified path with 2 criteria
        1.fileName
        2.file's extension that bind to that name
        """
        filePathList = []
        fileNameList = []
        extensionString = "|".join(extensionList)
        length = len(extensionList)
        count = 0
        for root, folders, files in os.walk(sourcePath):
            for file in files:
                hasDbName = IFile.has(f"(^.*{fileName}.*)", file)
                hasExtension = IFile.has(f".+\.({extensionString})", file)
                if hasDbName and hasExtension:
                    filePath = root + "\\" + file
                    filePathList.append(filePath)
                    fileNameList.append(file)
                    count += 1
                if count == length:
                    return filePathList, fileNameList

    @staticmethod
    def has(pattern, string):
        comObj = re.compile(pattern)
        matchObj = comObj.match(string)
        result = True
        try:
            matchObj.groups()
        except Exception:
            result = False
        finally:
            return result

    @staticmethod
    def parseToBoolean(char):
        """this method pareses the character that is in the user-excel file as 'True' or 'False'"""
        falsy = ("N", "n", "-")
        if char in falsy:
            return False
        else:
            return bool(char)

    @staticmethod
    def removeFile(filePath):
        result = True
        try:
            os.remove(filePath)
        except Exception:
            result = False
        finally:
            return result

    @staticmethod
    def splitFileExtension(fileName=None, fileNamePath=None): pass






if __name__ == "__main__":
    IFile.readUserInputFile("test_inputUserExcel.xlsx")