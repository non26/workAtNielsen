from openpyxl.styles.fills import PatternFill
from ..MarkOriginal import MarkOriginal
from ..ToLog import ToLog
import datetime
import os

class Week_MarkOriginal(MarkOriginal):

    def __init__(self, pathNADFile=None, sheetName="Sheet1"):
        MarkOriginal.__init__(self, pathNADFile, sheetName)

    def markSVNDWD(self, skipRow=None, col_sv = None
                   , col_nd=None, col_wd=None
                   , markSV=True, markND=True, markWD=True):
        print("sheetName", self.workSheet, markSV, markND, markWD, self.markColor.bgColor, skipRow, col_sv, col_nd, col_wd)
        if not markSV:
            # self.workSheet.cell(row=skipRow+4, column=columnSV+1).fill = yellowFill_NAD
            self.workSheet.cell(row=skipRow, column=col_sv + 1).fill = self.markColor
        if not markND:
            self.workSheet.cell(row=skipRow, column=col_nd + 1).fill = self.markColor
        if not markWD:
            self.workSheet.cell(row=skipRow, column=col_wd + 1).fill = self.markColor

class Week_ToLog(ToLog):

    def __init__(self, pathLogFile=None, sheets=None):
        ToLog.__init__(self, pathLogFile=pathLogFile, sheets=sheets)
        self.countRow = 2
        self.thisDay = datetime.datetime.now().strftime("%d/%m/%Y")

    def writeLogSVNDWD(self, chain=None
                       , pathNADFile=None, period=None, largest_MBD=None
                       , logSV=True, logND=True, logWD=True
                       , valueSV=None, valueND=None, valueWD=None
                       , dbname=None, report_name=None, report_line=None):
        fileNAD = os.path.basename(pathNADFile)
        if self.countRow == 2:
            self.workSheets[0].cell(row=1, column=1).value = "date(d/m/y)"
            self.workSheets[0].cell(row=1, column=2).value = "chain(service)"
            self.workSheets[0].cell(row=1, column=3).value = "fileNADName"
            self.workSheets[0].cell(row=1, column=4).value = "DBName"
            self.workSheets[0].cell(row=1, column=5).value = "period"
            self.workSheets[0].cell(row=1, column=6).value = "MBD"
            self.workSheets[0].cell(row=1, column=7).value = "Fact"
            self.workSheets[0].cell(row=1, column=8).value = "Result"
        if not logSV:
            self.workSheets[0].cell(row=self.countRow, column=1).value = self.thisDay
            self.workSheets[0].cell(row=self.countRow, column=2).value = chain
            self.workSheets[0].cell(row=self.countRow, column=3).value = fileNAD
            self.workSheets[0].cell(row=self.countRow, column=3).hyperlink = pathNADFile
            self.workSheets[0].cell(row=self.countRow, column=4).value = dbname
            self.workSheets[0].cell(row=self.countRow, column=5).value = period
            self.workSheets[0].cell(row=self.countRow, column=6).value = largest_MBD
            self.workSheets[0].cell(row=self.countRow, column=7).value = "SALE VALUE(trend-NAD)"
            self.workSheets[0].cell(row=self.countRow, column=8).value = "{0:,}".format(valueSV)
            self.workSheets[0].cell(row=self.countRow, column=8).fill = self.less if valueSV < 0 else self.more
            self.countRow += 1
        if not logND:
            self.workSheets[0].cell(row=self.countRow, column=1).value = self.thisDay
            self.workSheets[0].cell(row=self.countRow, column=2).value = chain
            self.workSheets[0].cell(row=self.countRow, column=3).value = fileNAD
            self.workSheets[0].cell(row=self.countRow, column=3).hyperlink = pathNADFile
            self.workSheets[0].cell(row=self.countRow, column=4).value = dbname
            self.workSheets[0].cell(row=self.countRow, column=5).value = period
            self.workSheets[0].cell(row=self.countRow, column=6).value = largest_MBD
            self.workSheets[0].cell(row=self.countRow, column=7).value = "ND_selling"
            self.workSheets[0].cell(row=self.countRow, column=8).value = valueND
            self.countRow += 1
        if not logWD:
            self.workSheets[0].cell(row=self.countRow, column=1).value = self.thisDay
            self.workSheets[0].cell(row=self.countRow, column=2).value = chain
            self.workSheets[0].cell(row=self.countRow, column=3).value = fileNAD
            self.workSheets[0].cell(row=self.countRow, column=3).hyperlink = pathNADFile
            self.workSheets[0].cell(row=self.countRow, column=4).value = dbname
            self.workSheets[0].cell(row=self.countRow, column=5).value = period
            self.workSheets[0].cell(row=self.countRow, column=6).value = largest_MBD
            self.workSheets[0].cell(row=self.countRow, column=7).value = "WD_selling"
            self.workSheets[0].cell(row=self.countRow, column=8).value = valueWD
            self.countRow += 1

    # def write_logNotFoundDBName(self, dict_notFoundDBName):
    #     self.workSheets[1].cell(row=1, column=1).value = "DBName"
    #     self.workSheets[1].cell(row=1, column=2).value = "Message"
    #     countRow = 2
    #     for keyDBName in dict_notFoundDBName:
    #         self.workSheets[1].cell(row=countRow, column=1).value = keyDBName
    #         self.workSheets[1].cell(row=countRow, column=2).value = dict_notFoundDBName[keyDBName]
    #         countRow += 1