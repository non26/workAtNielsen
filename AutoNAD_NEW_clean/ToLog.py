import openpyxl
from openpyxl.styles.fills import PatternFill
import os
import datetime
from .ConWorkBook import ConnWorkBook
class ToLog(ConnWorkBook):

    def __init__(self, pathLogFile=None, sheets=None):
        ConnWorkBook.__init__(self, pathFile=pathLogFile)
        self.pathLogFile = pathLogFile
        self.workSheets = []
        for sheet in sheets:
            self.workSheets.append(self.modifyWorkSheet(sheetName=sheet))
        self.more = PatternFill(patternType='solid', fgColor="80E1E1")
        self.less = PatternFill(patternType="solid", fgColor="E18080")

    def writeLogSVNDWD(self, chain=None
                       , pathNADFile=None, period=None, largest_MBD=None
                       , logSV=True, logND=True, logWD=True
                       , valueSV=None, valueND=None, valueWD=None
                       , dbname=None, report_name=None, report_line=None):
        pass

    def writeLogHidden(self, failHidden=None, pathNADFile=None, report_name=None):
        pass

    def write_logNotFoundDBName(self, dict_notFoundDBName, orderOfSheet=0):
        self.workSheets[orderOfSheet].cell(row=1, column=1).value = "DBName"
        self.workSheets[orderOfSheet].cell(row=1, column=2).value = "Message"
        countRow = 2
        for keyDBName in dict_notFoundDBName:
            self.workSheets[orderOfSheet].cell(row=countRow, column=1).value = keyDBName
            self.workSheets[orderOfSheet].cell(row=countRow, column=2).value = dict_notFoundDBName[keyDBName]
            countRow += 1