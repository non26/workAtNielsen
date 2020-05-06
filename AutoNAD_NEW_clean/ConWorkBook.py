import openpyxl
class ConnWorkBook():
    def __init__(self, pathFile):
        """
        :param pathFile:
        """
        self.workSheet = None
        self.pathFile = pathFile
        try:
            self.workBook = openpyxl.load_workbook(self.pathFile)
        except (Exception,): # if work book does't exist then create one
            self.workBook = openpyxl.workbook.Workbook()

    @property
    def workSheetAccessor(self):
        return self.workSheet

    @workSheetAccessor.setter
    def workSheetAccessor(self, another):
        self.workSheet = self.workBook[another]

    @workSheetAccessor.deleter
    def workSheetAccessor(self):
        del self.workSheet

    @property
    def workBookAccessor(self):
        return self.workBook

    @workBookAccessor.setter
    def workBookAccessor(self, another):
        self.workBook = openpyxl.load_workbook(another)

    @workBookAccessor.deleter
    def workBookAccessor(self):
        del self.workBook

    def selectWorkSheet(self, sheetName="Sheet1"):
        """
        :param sheetName:
        :return:
        """
        try:
            self.workSheet = self.workBook[sheetName]
        except Exception : # If work sheet does't exist then create it.
            self.workSheet = self.workBook.create_sheet(title=sheetName)

    def modifyWorkSheet(self, sheetName="Sheet1"):
        """
        :param sheetName:
        :return:
        """
        self.workSheet = sheetName
        if sheetName in self.workBook.sheetnames:
            self.workBook.remove(self.workBook[sheetName])
            self.workSheet = self.workBook.create_sheet(sheetName)
        else:
            self.workSheet = self.workBook.create_sheet(sheetName)
        return self.workSheet

    def saveAction(self):
        self.workBook.save(self.pathFile)

    def closeWorkBook(self):
        self.workBook.close()
